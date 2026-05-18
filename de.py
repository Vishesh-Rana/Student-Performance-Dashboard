import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.figure_factory as ff
import plotly.graph_objects as go
import numpy as np
import base64
import os
import io
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import json
from openpyxl import load_workbook  # Added for parsing

st.set_page_config(layout="wide", page_title="Student Performance Analysis Dashboard")

# ---- Google Drive Setup ----
@st.cache_resource
def initialize_drive_service():
    """Initialize Google Drive service using service account credentials"""
    try:
        credentials_info = st.secrets["google_service_account"]
        credentials = Credentials.from_service_account_info(
            credentials_info,
            scopes=['https://www.googleapis.com/auth/drive.readonly']
        )
        service = build('drive', 'v3', credentials=credentials)
        return service
    except Exception as e:
        st.error(f"Failed to initialize Google Drive service: {str(e)}")
        return None

@st.cache_data(ttl=3600)  # Cache for 1 hour
def download_file_from_drive(_service, file_id, file_name):
    """Download file from Google Drive and return as bytes"""
    try:
        request = _service.files().get_media(fileId=file_id)
        file_io = io.BytesIO()
        downloader = MediaIoBaseDownload(file_io, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        file_io.seek(0)
        return file_io.getvalue()
    except Exception as e:
        st.error(f"Error downloading {file_name}: {str(e)}")
        return None

@st.cache_data(ttl=3600)  # Cache for 1 hour
def load_excel_from_drive(_service, file_id, file_name):
    """Load Excel file from Google Drive into pandas DataFrame"""
    file_content = download_file_from_drive(_service, file_id, file_name)
    if file_content:
        return pd.read_excel(io.BytesIO(file_content), sheet_name=None, engine='openpyxl')
    return None

@st.cache_data(ttl=3600)  # Cache for 1 hour
def load_image_from_drive(_service, file_id, file_name):
    """Load image file from Google Drive and convert to base64"""
    file_content = download_file_from_drive(_service, file_id, file_name)
    if file_content:
        return base64.b64encode(file_content).decode()
    return None

@st.cache_data(ttl=3600)
def load_workbook_from_drive(_service, file_id, file_name):

    file_content = download_file_from_drive(
        _service,
        file_id,
        file_name
    )

    if not file_content:
        return None

    wb = load_workbook(
        filename=io.BytesIO(file_content),
        data_only=True
    )


    return wb

# ---- Custom CSS ----
st.markdown("""
    <style>
    .metric-header {
        background-color: #87cefa;
        color: black;
        text-align: center;
        font-weight: bold;
        font-size: 0.9em;
        border-radius: 5px 5px 0 0;
        padding: 4px 0 4px 0;
        margin-bottom: 0px;
    }
    .metric-card {
        background: white;
        border-radius: 0 0 5px 5px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.04);
        padding: 6px 0 3px 0;
        margin-bottom: 8px;
        text-align: center;
    }
    .metric-value {
        font-size: 1.5em;
        font-weight: 600;
        color: #222;
    }
    .metric-label {
        font-size: 0.8em;
        color: #666;
    }
    .block-container {
        padding-top: 1rem;
        padding-bottom: 0rem;
    }
    .main-header {
        position: sticky;
        top: 0;
        z-index: 999;
        background-color: white;
        padding: 10px 0;
        margin-bottom: 20px;
    }
    </style>
""", unsafe_allow_html=True)


# ==========================================
# ---- NEW: Tertiary Parsing Functions -----
# ==========================================

def parse_excel_blocks_with_empty_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    results = []

    # STEP 1: detect YEAR blocks
    header_row = None
    for r in rows:
        if r and any(isinstance(v, str) and "YEAR" in v.upper() for v in r):
            header_row = r
            break

    if not header_row:
        return pd.DataFrame()

    block_starts = []
    for i, v in enumerate(header_row):
        if isinstance(v, str) and "YEAR" in v.upper():
            block_starts.append((i, v.strip()))

    block_width = 3  # fixed 3-column structure

    # STEP 2: process each YEAR block
    for start_col, year in block_starts:
        current_semester = None
        table_started = False
        row_found_in_table = False

        for r in rows:
            if not r or start_col >= len(r):
                continue

            cell = r[start_col] if start_col < len(r) else None

            # SEMESTER
            if isinstance(cell, str) and "SEMESTER" in cell.upper():
                if table_started and not row_found_in_table:
                    results.append({
                        "Year": year,
                        "Semester": current_semester,
                        "Code": None,
                        "Unit Name": None,
                        "Grade": None
                    })
                current_semester = cell.strip()
                table_started = False
                row_found_in_table = False
                continue

            # HEADER
            if cell == "CODE":
                table_started = True
                row_found_in_table = False
                continue

            # SKIP IF NOT IN TABLE
            if not table_started:
                continue

            # EMPTY ROW
            if all(r[c] is None or str(r[c]).strip() == "" for c in range(start_col, min(start_col + block_width, len(r)))):
                continue

            # EXTRACT
            code = r[start_col]
            unit = r[start_col + 1] if start_col + 1 < len(r) else None
            grade = r[start_col + 2] if start_col + 2 < len(r) else None

            if isinstance(unit, str) and unit.strip():
                results.append({
                    "Year": year,
                    "Semester": current_semester,
                    "Code": code,
                    "Unit Name": unit,
                    "Grade": grade
                })
                row_found_in_table = True

        # flush last semester block
        if table_started and not row_found_in_table:
            results.append({
                "Year": year,
                "Semester": current_semester,
                "Code": None,
                "Unit Name": None,
                "Grade": None
            })

    return pd.DataFrame(results)

def process_multiple_sheets(_service, file_id, file_name):
    
    wb = load_workbook_from_drive(_service, file_id, file_name)

    if wb is None:
        return pd.DataFrame()

    ws1 = wb['Tertiary Student List']
    data = list(ws1.values)
    df_temp = pd.DataFrame(data[1:], columns=data[0])
    df_temp = df_temp[['Name', 'School', 'Course']]
    
    remove_vals = ['Tertiary Student List', 'Template']
    sheet_list = [x for x in list(wb.sheetnames) if x not in remove_vals]

    all_dfs = []
    
    for sheet in sheet_list:
        ws = wb[sheet]


        df = parse_excel_blocks_with_empty_rows(ws)

        if not df.empty:
            df["Student_Name"] = sheet  # add sheet name as student
            
            # Step 0: Normalize to uppercase (Added missing .str. to avoid SyntaxError)
            df['Grade'] = df['Grade'].str.strip().str.upper()
            df['Grade'] = df['Grade'].fillna('MISSING')  # Mark missing grades explicitly

            # Step 1: Create result column
            df['Result'] = df['Grade'].apply(lambda x: 'FAIL' if x in ['E', 'F', 'FAIL']
                                             else 'MISSING' if x in ['MISSING']
                                             else 'SPECIAL' if x in ['SPECIAL']
                                                    else 'PASS')

            # Step 2: Replace PASS/FAIL in Grade with "NOT AVAILABLE"
            df.loc[df['Grade'].isin(['PASS', 'FAIL','MISSING', 'SPECIAL']), 'Grade'] = 'NOT AVAILABLE'
            
            all_dfs.append(df)

    if all_dfs:
        final_df = pd.concat(all_dfs, ignore_index=True)
    else:
        final_df = pd.DataFrame()

    tertiary_df = final_df.merge(df_temp, left_on='Student_Name', right_on='Name', how='left').drop(columns=['Name'])
    tertiary_df = tertiary_df[['Student_Name', 'School', 'Course', 'Year', 'Semester', 'Code', 'Unit Name', 'Grade', 'Result']]


    return tertiary_df


# ---- Load Data from Google Drive ----
def load_data():
    """Load all data from Google Drive and local files"""
    service = initialize_drive_service()
    if not service:
        st.error("Cannot connect to Google Drive. Please check your service account configuration.")
        st.stop()
    
    try:
        # Get file IDs from secrets
        file_ids = st.secrets["google_drive_files"]

        # Load team result files (From Google Drive)
        files_and_teams = [
            (file_ids["team_kathy"], "Team Kathy"),
            (file_ids["team_kelly"], "Team Kelly"),
            (file_ids["team_lissette"], "Team Lissette"),
        ]

        dfs = []
        for file_id, team in files_and_teams:
            try:
                all_sheets = load_excel_from_drive(service, file_id, f"{team} Results")
                if all_sheets:
                    for sheet_name, df in all_sheets.items():
                        df["Team Name"] = team
                        # Replace NA, N/A, and similar values with "Not Appeared" across all columns
                        df = df.replace({
                            'NA': 'Not Appeared', 'N/A': 'Not Appeared', 'N.A': 'Not Appeared',
                            'n/a': 'Not Appeared', 'na': 'Not Appeared', 'n.a': 'Not Appeared',
                            'N.A.': 'Not Appeared', 'N/A/': 'Not Appeared'
                        })
                        dfs.append(df)
                else:
                    st.warning(f"Could not load data for {team}")
            except Exception as e:
                st.error(f"Error loading {team} data: {str(e)}")

        if not dfs:
            st.error("No team data could be loaded.")
            st.stop()

        df_main = pd.concat(dfs, ignore_index=True)

        # Load High School Data Sheet (From Google Drive)
        high_school_file_id = file_ids.get("high_school_data", "")
        high_school_unique_students = None
        if high_school_file_id:
            high_school_data = load_excel_from_drive(service, high_school_file_id, "High School Data")
            if high_school_data:
                high_school_df = list(high_school_data.values())[0]
                high_school_df = high_school_df.rename(columns={"Name": "Student"})
                high_school_df = high_school_df.replace({
                    'NA': 'Not Appeared', 'N/A': 'Not Appeared', 'N.A': 'Not Appeared',
                    'n/a': 'Not Appeared', 'na': 'Not Appeared', 'n.a': 'Not Appeared',
                    'N.A.': 'Not Appeared', 'N/A/': 'Not Appeared'
                })
                high_school_unique_students = high_school_df["Student"].dropna().astype(str).str.strip().nunique()
                df_main = df_main.merge(high_school_df, how="left", on="Student")
            else:
                st.warning("Could not load High School Data Sheet")
        else:
            st.info("High School Data Sheet not configured - using team data only")

        # Load Dropout Data (From Google Drive)
        dropout_file_id = file_ids.get("dropout_data", "")
        dropout_df = None
        if dropout_file_id:
            dropout_excel = load_excel_from_drive(service, dropout_file_id, "Dropout Data")
            if dropout_excel:
                sheet_name = list(dropout_excel.keys())[0]
                dropout_df = dropout_excel[sheet_name]

        # ==========================================
        # ---- NEW: Load Tertiary Data LOCALLY -----
        # ==========================================
        
        # Ensure this matches the name of your raw unparsed file
        tertiary_file_id = file_ids.get("tertiary_data", "")
        
        tertiary_df = None
        if tertiary_file_id:
            try:
                # 1. Parse the multiple sheets to get the main DF
                tertiary_df = process_multiple_sheets(service, tertiary_file_id, "Tertiary Restults")
                
                # 3. Rename 'Student_Name' to 'Student' to perfectly match Tab 5 code requirements
                if "Student_Name" in tertiary_df.columns:
                    tertiary_df = tertiary_df.rename(columns={"Student_Name": "Student"})

            except Exception as e:
                st.error(f"Error parsing theTertiary Excel file: {str(e)}")
        else:
            st.warning(f"Unable to load Tertiary data from drive. Please check the configurations.")

        # Return all main dataframes
        return df_main, high_school_unique_students, dropout_df, tertiary_df

    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        st.stop()

# Function to load logo from local file
def get_logo_base64():
    """Load logo from local file and convert to base64"""
    try:
        logo_path = "SAM Elimu Logo-white_edited.png"
        with open(logo_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode()
    except Exception as e:
        # Return none silently if no logo found locally
        return None

# Load data and logo
with st.spinner("Loading data from Google Drive..."):
    df_main, high_school_unique_students, dropout_df, tertiary_df = load_data()

# Load logo from local file
logo_base64 = get_logo_base64()

# ---- Clean up duplicate columns after merge ----
if "Form_x" in df_main.columns or "Form_y" in df_main.columns:
    df_main["Form"] = df_main.get("Form_x", pd.Series(dtype=object)).combine_first(df_main.get("Form_y", pd.Series(dtype=object)))
    df_main = df_main.drop(columns=[col for col in ["Form_x", "Form_y"] if col in df_main.columns])
if "School_x" in df_main.columns or "School_y" in df_main.columns:
    df_main["School"] = df_main.get("School_x", pd.Series(dtype=object)).combine_first(df_main.get("School_y", pd.Series(dtype=object)))
    df_main = df_main.drop(columns=[col for col in ["School_x", "School_y"] if col in df_main.columns])

# ---- Data Cleaning ----
if "School" in df_main.columns and "Student" in df_main.columns:
    df_main = df_main[~(df_main["School"].isna() & df_main["Student"].isna())]
    df_main = df_main[~((df_main["School"].astype(str).str.strip() == "") & (df_main["Student"].astype(str).str.strip() == ""))]
elif "Student" in df_main.columns:
    df_main = df_main[~(df_main["Student"].isna())]
    df_main = df_main[~(df_main["Student"].astype(str).str.strip() == "")]

subject_columns = [
    "Maths", "English", "Kiswahili", "Chemistry", "Biology", "Physics", "CRE", "Geography",
    "History", "Agriculture", "Business Studies", "French", "Computer studies", "Home Science",
    "Woodwork"
]

def all_subjects_empty(row):
    found_data = False
    for col in subject_columns:
        val = row.get(col, np.nan)
        if pd.notna(val):
            sval = str(val).strip()
            if sval != "" and sval != "Not Appeared":
                found_data = True
                break
    return not found_data

df_main = df_main[~df_main.apply(all_subjects_empty, axis=1)].reset_index(drop=True)

# ---- Calculate M% (Overall Percentage) from Subject Scores ----
def calculate_m_percentage(row):
    valid_scores = []
    for col in subject_columns:
        if col in row.index:
            val = row[col]
            if pd.notna(val):
                sval = str(val).strip()
                if sval != "" and sval != "Not Appeared":
                    try:
                        numeric_val = float(sval)
                        if 0 <= numeric_val <= 100:
                            valid_scores.append(numeric_val)
                    except (ValueError, TypeError):
                        continue
    if len(valid_scores) > 0:
        return round(sum(valid_scores) / len(valid_scores), 2)
    else:
        return 0.0

df_main["M%"] = df_main.apply(calculate_m_percentage, axis=1)

# ---- Add Remark Column Based on Mean Grade ----
def grade_to_remark(grade):
    if pd.isna(grade):
        return "Unknown"
    grade = str(grade).strip().upper()
    if grade in ["B", "B+", "A-", "A"]:
        return "Exceeding Expectation"
    elif grade in ["C+", "B-"]:
        return "Meeting Expectation"
    elif grade in ["C", "C-", "D+", "D", "D-", "E"]:
        return "Below Expectation"
    else:
        return "Unknown"

if "Mean Grade" in df_main.columns:
    df_main["Remark"] = df_main["Mean Grade"].apply(grade_to_remark)

# ---- Page Title ----
if logo_base64:
    st.markdown(f"""
        <div class="main-header">
            <div style='background-color: #FFC300; padding: 8px; border-radius: 8px; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.1); position: relative; display: flex; align-items: center; justify-content: space-between;'>
                <h1 style='color: black; margin: 0; font-size: 1.6em; font-weight: bold; flex: 1; text-align: center;'>Student Performance Analysis Dashboard</h1>
                <img src="data:image/png;base64,{logo_base64}" style="height: 80px; width: auto; margin: 10px; padding-top: 20px" alt="SAM Elimu Logo">
            </div>
        </div>
    """, unsafe_allow_html=True)
else:
    st.markdown("""
        <div class="main-header">
            <div style='background-color: #FFC300; padding: 8px; border-radius: 8px; text-align: center; box-shadow: 0 2px 4px rgba(0,0,0,0.1); position: relative;'>
                <h1 style='color: black; margin: 0; font-size: 1.6em; font-weight: bold;'>Student Performance Analysis Dashboard</h1>
            </div>
        </div>
    """, unsafe_allow_html=True)

# ---- Tab Structure ----
tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Overall Analysis", "👨‍🎓 Student Analysis", "📋 Detailed Data", "🚪 Dropouts", "🎓 Tertiary Analysis"])

# --- Tab 1 ---
with tab1:
    main_col, filter_col = st.columns([4, 1])

    with filter_col:
        st.markdown("""
            <div style='background-color: #87cefa; padding: 1px 0; border-radius: 1px; text-align: center; margin-bottom: 0.1px;'>
                <span style='color: black; font-weight: bold; font-size: 16px;'>🔍 Filter Students</span>
            </div>
        """, unsafe_allow_html=True)
        
        team = st.selectbox("Team Name", options=["All"] + sorted([str(x) for x in df_main["Team Name"].dropna().unique().tolist()]))
        
        filtered_for_options = df_main.copy()
        if team and team != "All":
            filtered_for_options = filtered_for_options[filtered_for_options["Team Name"].astype(str) == team]
        
        if "Form" in filtered_for_options.columns:
            available_forms = sorted([str(x) for x in filtered_for_options["Form"].dropna().unique().tolist()])
            form = st.multiselect("Form", options=available_forms)
            if form:
                filtered_for_options = filtered_for_options[filtered_for_options["Form"].astype(str).isin(form)]
        else:
            form = []
        
        available_periods = sorted([str(x) for x in filtered_for_options["Period"].dropna().unique().tolist()])
        period = st.multiselect("Period (type to search)", options=available_periods, max_selections=5, help="Start typing to quickly find a period.")
        
        if period:
            filtered_for_options = filtered_for_options[filtered_for_options["Period"].astype(str).isin(period)]
        
        if "School" in filtered_for_options.columns:
            available_schools = sorted([str(x) for x in filtered_for_options["School"].dropna().unique().tolist()])
            school = st.multiselect("School", options=available_schools)
            if school:
                filtered_for_options = filtered_for_options[filtered_for_options["School"].astype(str).isin(school)]
        else:
            school = []
        
        if "Mean Grade" in filtered_for_options.columns:
            available_grades = sorted([str(x) for x in filtered_for_options["Mean Grade"].dropna().unique().tolist()])
            grade = st.multiselect("Mean Grade", options=available_grades)
            if grade:
                filtered_for_options = filtered_for_options[filtered_for_options["Mean Grade"].astype(str).isin(grade)]
        else:
            grade = []
        
        if "Donor" in filtered_for_options.columns:
            available_donors = sorted([str(x) for x in filtered_for_options["Donor"].dropna().unique().tolist()])
            donor = st.multiselect("Donor", options=available_donors)
            if donor:
                filtered_for_options = filtered_for_options[filtered_for_options["Donor"].astype(str).isin(donor)]
        else:
            donor = []
        
        if "Home County" in filtered_for_options.columns:
            available_counties = sorted([str(x) for x in filtered_for_options["Home County"].dropna().unique().tolist()])
            county = st.multiselect("Home County", options=available_counties)
        else:
            county = []
        
        marks_range = st.slider("% Marks", 0, 100, (0, 100))

    filtered = df_main.copy()
    
    if team and team != "All":
        filtered = filtered[filtered["Team Name"].astype(str) == team]
    if form:
        filtered = filtered[filtered["Form"].astype(str).isin(form)]
    if period:
        filtered = filtered[filtered["Period"].astype(str).isin(period)]
    if school:
        filtered = filtered[filtered["School"].astype(str).isin(school)]
    if grade:
        filtered = filtered[filtered["Mean Grade"].astype(str).isin(grade)]
    if donor:
        filtered = filtered[filtered["Donor"].astype(str).isin(donor)]
    if county:
        filtered = filtered[filtered["Home County"].astype(str).isin(county)]
    if "M%" in filtered.columns:
        filtered = filtered[(filtered["M%"] >= marks_range[0]) & (filtered["M%"] <= marks_range[1])]
    
    filtered = filtered.copy()

    for col in subject_columns:
        if col in filtered.columns:
            filtered[col] = filtered[col].replace("Not Appeared", np.nan)
            filtered[col] = pd.to_numeric(filtered[col], errors='coerce')

    with main_col:
        st.markdown("---")
        main_cols_row1 = st.columns(3)

        with main_cols_row1[0]:
            st.markdown('<div class="metric-header">Number of Students</div>', unsafe_allow_html=True)
            unique_students = filtered["Student"].nunique() if "Student" in filtered.columns else 0
            hs_students = high_school_unique_students if high_school_unique_students is not None else "N/A"
            st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-value">{hs_students}</div>
                </div>
            """, unsafe_allow_html=True)

        with main_cols_row1[1]:
            st.markdown('<div class="metric-header">Average Score in Sciences</div>', unsafe_allow_html=True)
            science_subjects = ["Maths", "Biology", "Chemistry", "Physics"]
            science_metrics = []
            for subject in science_subjects:
                if subject in filtered.columns:
                    science_metrics.append((subject, filtered[subject].mean()))
            
            if science_metrics:
                sci_cols = st.columns(len(science_metrics))
                for i, (label, value) in enumerate(science_metrics):
                    display_value = f"{value:.0f}" if pd.notnull(value) else "--"
                    with sci_cols[i]:
                        st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-label">{label}</div>
                                <div class="metric-value">{display_value}</div>
                            </div>
                        """, unsafe_allow_html=True)

        with main_cols_row1[2]:
            st.markdown('<div class="metric-header">Average Score in Languages</div>', unsafe_allow_html=True)
            language_subjects = ["English", "Kiswahili", "French"]
            language_metrics = []
            for subject in language_subjects:
                if subject in filtered.columns:
                    language_metrics.append((subject, filtered[subject].mean()))
            
            if language_metrics:
                lang_cols = st.columns(len(language_metrics))
                for i, (label, value) in enumerate(language_metrics):
                    display_value = f"{value:.0f}" if pd.notnull(value) else "--"
                    with lang_cols[i]:
                        st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-label">{label}</div>
                                <div class="metric-value">{display_value}</div>
                            </div>
                        """, unsafe_allow_html=True)

        main_cols_row2 = st.columns([1, 2])

        with main_cols_row2[0]:
            st.markdown('<div class="metric-header">Average Score in Humanities</div>', unsafe_allow_html=True)
            humanities_subjects = ["History", "Geography", "CRE"]
            humanities_metrics = []
            for subject in humanities_subjects:
                if subject in filtered.columns:
                    humanities_metrics.append((subject, filtered[subject].mean()))
            
            if humanities_metrics:
                hum_cols = st.columns(len(humanities_metrics))
                for i, (label, value) in enumerate(humanities_metrics):
                    display_value = f"{value:.0f}" if pd.notnull(value) else "--"
                    with hum_cols[i]:
                        st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-label">{label}</div>
                                <div class="metric-value">{display_value}</div>
                            </div>
                        """, unsafe_allow_html=True)

        with main_cols_row2[1]:
            st.markdown('<div class="metric-header">Average Score in Technical Subjects</div>', unsafe_allow_html=True)
            technical_subjects = ["Computer studies", "Business Studies", "Woodwork", "Home Science", "Agriculture"]
            technical_metrics = []
            for subject in technical_subjects:
                if subject in filtered.columns:
                    technical_metrics.append((subject, filtered[subject].mean()))
            
            if technical_metrics:
                tech_cols = st.columns(len(technical_metrics))
                for i, (label, value) in enumerate(technical_metrics):
                    display_value = f"{value:.0f}" if pd.notnull(value) else "--"
                    with tech_cols[i]:
                        st.markdown(f"""
                            <div class="metric-card">
                                <div class="metric-label">{label}</div>
                                <div class="metric-value">{display_value}</div>
                            </div>
                        """, unsafe_allow_html=True)

    # --- FULL WIDTH CHARTS ---
    st.markdown("---")
    chart1, chart2 = st.columns(2)

    if "Remark" in filtered.columns:
        remark_counts = filtered["Remark"].value_counts()
        fig1 = px.pie(
            values=remark_counts.values,
            names=remark_counts.index,
            title="Performance Level Distribution",
            color_discrete_sequence=px.colors.qualitative.Set3
        )
        fig1.update_traces(hovertemplate='<b>%{label}</b><br>Count: %{value}<br>Percentage: %{percent}<extra></extra>')
        chart1.plotly_chart(fig1, use_container_width=True)

    if "Mean Grade" in filtered.columns:
        grade_counts = filtered["Mean Grade"].value_counts()
        if len(grade_counts) > 0:
            grade_order = ["A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "E"]
            ordered_grades = [grade for grade in grade_order if grade in grade_counts.index]
            ordered_counts = [grade_counts[grade] for grade in ordered_grades]
            
            fig2 = px.bar(
                x=ordered_grades,
                y=ordered_counts,
                labels={"x": "Grade", "y": "Number of Students"},
                title="Student Distribution by Grade",
                color=ordered_counts,
                color_continuous_scale="viridis"
            )
            fig2.update_traces(hovertemplate='<b>Grade %{x}</b><br>Students: %{y}<extra></extra>')
            fig2.update_layout(
                xaxis={'categoryorder': 'array', 'categoryarray': ordered_grades},
                showlegend=False
            )
            chart2.plotly_chart(fig2, use_container_width=True)
        else:
            chart2.info("No grade data available for this selection.")

    chart3, chart4 = st.columns(2)
    existing_subjects = [sub for sub in subject_columns if sub in filtered.columns]
    if existing_subjects:
        subject_avg = filtered[existing_subjects].mean().sort_values()
        concern_subjects = subject_avg[subject_avg < 55]
        if not concern_subjects.empty:
            fig3 = px.bar(
                x=concern_subjects.index,
                y=concern_subjects.values,
                labels={"x": "Subject", "y": "Average Score (%)"},
                title="Subjects Needing Attention (Avg < 55%)",
                color=concern_subjects.values,
                color_continuous_scale="Reds"
            )
            fig3.update_traces(hovertemplate='<b>%{x}</b><br>Average: %{y:.1f}%<extra></extra>')
            fig3.update_layout(showlegend=False, xaxis_tickangle=-45)
            chart3.plotly_chart(fig3, use_container_width=True)
        else:
            chart3.info("No subjects of concern (all averages >= 55%).")

    if "M%" in filtered.columns and "Student" in filtered.columns:
        top_students = filtered.sort_values("M%", ascending=False).drop_duplicates("Student").head(5)
        fig4 = px.bar(
            top_students,
            x="Student",
            y="M%",
            title="Overall Performance Distribution",
            color="M%",
            color_continuous_scale="Greens"
        )
        fig4.update_layout(
            xaxis_title="Student",
            yaxis_title="M%",
            showlegend=False
        )
        chart4.plotly_chart(fig4, use_container_width=True)

# --- Tab 2 ---
with tab2:
    st.markdown("### 👨‍🎓 Individual Student Analysis")
    if "Student" in df_main.columns:
        student_list = []
        for student in df_main["Student"].dropna().unique():
            student_str = str(student).strip()
            if (student_str and 
                student_str not in ["Category Distribution", "CATEGORY DISTRIBUTION", "category distribution"] and
                not student_str.lower().startswith("category") and
                not student_str.lower().startswith("total") and
                not student_str.lower().startswith("average") and
                len(student_str) > 2):
                student_list.append(student_str)
        student_list = sorted(student_list)
        selected_student = st.selectbox("Select a Student", options=student_list)
        if selected_student:
            student_data = df_main[df_main["Student"] == selected_student]
            if not student_data.empty:
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("#### 📊 Student Information")
                    if "School" in student_data.columns:
                        st.write(f"**School:** {student_data['School'].iloc[0] if not pd.isna(student_data['School'].iloc[0]) else 'N/A'}")
                    if "Form" in student_data.columns:
                        st.write(f"**Form:** {student_data['Form'].iloc[0] if not pd.isna(student_data['Form'].iloc[0]) else 'N/A'}")
                    if "Team Name" in student_data.columns:
                        st.write(f"**Team:** {student_data['Team Name'].iloc[0] if not pd.isna(student_data['Team Name'].iloc[0]) else 'N/A'}")
                    if "Home County" in student_data.columns:
                        st.write(f"**Home County:** {student_data['Home County'].iloc[0] if not pd.isna(student_data['Home County'].iloc[0]) else 'N/A'}")
                with col2:
                    st.markdown("#### 🎯 Performance Metrics")
                    if "Mean Grade" in student_data.columns:
                        st.write(f"**Mean Grade:** {student_data['Mean Grade'].iloc[0] if not pd.isna(student_data['Mean Grade'].iloc[0]) else 'N/A'}")
                    if "M%" in student_data.columns:
                        st.write(f"**Overall Percentage:** {student_data['M%'].iloc[0] if not pd.isna(student_data['M%'].iloc[0]) else 'N/A'}%")
                    if "Remark" in student_data.columns:
                        st.write(f"**Remark:** {student_data['Remark'].iloc[0] if not pd.isna(student_data['Remark'].iloc[0]) else 'N/A'}")
                
                available_periods = sorted([str(x) for x in student_data["Period"].dropna().unique()]) if "Period" in student_data.columns else []
                selected_period = None
                if available_periods:
                    selected_period = st.selectbox("Select Period for Subject Performance", options=available_periods)
                st.markdown("#### 📚 Subject Performance")
                if selected_period:
                    period_data = student_data[student_data["Period"].astype(str) == selected_period]
                else:
                    period_data = student_data
                subject_scores = []
                subject_names = []
                for subject in subject_columns:
                    if subject in period_data.columns:
                        score = period_data[subject].iloc[0]
                        if pd.notna(score) and str(score).strip() not in ["Not Appeared", ""]:
                            try:
                                numeric_score = float(score)
                                subject_scores.append(numeric_score)
                                subject_names.append(subject)
                            except:
                                pass
                if subject_scores and subject_names:
                    fig_subjects = px.bar(
                        x=subject_names,
                        y=subject_scores,
                        title=f"Subject Scores for {selected_student} ({selected_period if selected_period else 'All Periods'})",
                        labels={"x": "Subject", "y": "Score"},
                        color=subject_scores,
                        color_continuous_scale="viridis"
                    )
                    fig_subjects.update_traces(hovertemplate='<b>%{x}</b><br>Subject Score: %{y}<extra></extra>')
                    fig_subjects.add_hline(y=60, line_dash="dash", line_color="red", annotation_text="Pass Mark (60%)")
                    st.plotly_chart(fig_subjects, use_container_width=True)
                    avg_score = np.mean(subject_scores)
                    subjects_below_60 = [name for name, score in zip(subject_names, subject_scores) if score < 60]
                    subjects_above_80 = [name for name, score in zip(subject_names, subject_scores) if score >= 80]
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Average Score", f"{avg_score:.1f}%")
                    with col2:
                        st.metric("Subjects Below 60%", len(subjects_below_60))
                    with col3:
                        st.metric("Subjects Above 80%", len(subjects_above_80))
                    if subjects_below_60:
                        st.warning(f"**Subjects needing improvement:** {', '.join(subjects_below_60)}")
                    if subjects_above_80:
                        st.success(f"**Strong subjects:** {', '.join(subjects_above_80)}")
                
                not_appeared_subjects = []
                for subject in subject_columns:
                    if subject in period_data.columns:
                        score = period_data[subject].iloc[0]
                        if pd.notna(score) and str(score).strip() == "Not Appeared":
                            not_appeared_subjects.append(subject)
                if not_appeared_subjects:
                    st.info(f"**Subjects not appeared:** {', '.join(not_appeared_subjects)}")
                
                st.markdown("#### 📈 Progress Over Time")
                student_all_periods = df_main[df_main["Student"] == selected_student]
                
                if "Period" in student_all_periods.columns:
                    unique_periods = student_all_periods["Period"].dropna().unique()
                    def period_to_float(period_str):
                        try:
                            return float(str(period_str).strip())
                        except:
                            return 0.0
                    sorted_periods = sorted(unique_periods, key=period_to_float)
                    
                    if len(unique_periods) > 1:
                        progress_data = []
                        for period in sorted_periods:
                            period_data = student_all_periods[student_all_periods["Period"] == period]
                            if not period_data.empty:
                                latest_record = period_data.iloc[-1]
                                row_data = {"Period": str(period)}
                                if "M%" in period_data.columns and pd.notna(latest_record["M%"]):
                                    try:
                                        m_percent = float(latest_record["M%"])
                                        if m_percent > 0 and m_percent <= 100:
                                            row_data["Overall %"] = m_percent
                                    except:
                                        pass
                                for subject in subject_columns:
                                    if subject in period_data.columns and pd.notna(latest_record[subject]):
                                        try:
                                            score_val = str(latest_record[subject]).strip()
                                            if score_val != "Not Appeared" and score_val != "":
                                                score = float(score_val)
                                                if score > 0 and score <= 100:
                                                    row_data[subject] = score
                                        except:
                                            pass
                                if len(row_data) > 1:
                                    progress_data.append(row_data)
                        
                        if len(progress_data) > 1:
                            progress_df = pd.DataFrame(progress_data)
                            if "Overall %" in progress_df.columns and progress_df["Overall %"].notna().sum() > 1:
                                overall_df = progress_df.dropna(subset=["Overall %"])
                                if len(overall_df) > 1:
                                    fig_overall = px.line(
                                        overall_df,
                                        x="Period",
                                        y="Overall %",
                                        title=f"Overall Performance Trend for {selected_student}",
                                        markers=True,
                                        line_shape="linear"
                                    )
                                    fig_overall.update_layout(
                                        xaxis_title="Period",
                                        yaxis_title="Overall Percentage (%)",
                                        xaxis=dict(type='category'),
                                        showlegend=True
                                    )
                                    st.plotly_chart(fig_overall, use_container_width=True)
                                    
                                    first_score = overall_df["Overall %"].iloc[0]
                                    last_score = overall_df["Overall %"].iloc[-1]
                                    change = last_score - first_score
                                    
                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        st.metric("First Period", f"{first_score:.1f}%")
                                    with col2:
                                        st.metric("Latest Period", f"{last_score:.1f}%")
                                    with col3:
                                        st.metric("Change", f"{change:+.1f}%", delta=f"{change:+.1f}%")
                                else:
                                    st.info("Overall percentage data available but insufficient valid data points for trend.")
                            else:
                                st.info("Overall percentage data not available or insufficient for trend analysis.")
                            
                            subject_cols = [col for col in progress_df.columns if col not in ["Period", "Overall %"]]
                            if subject_cols:
                                melted_data = []
                                for _, row in progress_df.iterrows():
                                    for subject in subject_cols:
                                        if pd.notna(row[subject]):
                                            melted_data.append({
                                                "Period": row["Period"],
                                                "Subject": subject,
                                                "Score": row[subject]
                                            })
                                if melted_data:
                                    melted_df = pd.DataFrame(melted_data)
                                    subject_counts = melted_df.groupby("Subject").size()
                                    valid_subjects = subject_counts[subject_counts >= 2].index.tolist()
                                    if valid_subjects:
                                        filtered_melted = melted_df[melted_df["Subject"].isin(valid_subjects)]
                                        fig_subjects = px.line(
                                            filtered_melted,
                                            x="Period",
                                            y="Score",
                                            color="Subject",
                                            title=f"Subject-wise Performance Trend for {selected_student}",
                                            markers=True
                                        )
                                        fig_subjects.update_layout(
                                            xaxis_title="Period",
                                            yaxis_title="Score (%)",
                                            xaxis=dict(type='category'),
                                            showlegend=True
                                        )
                                        st.plotly_chart(fig_subjects, use_container_width=True)
                                    else:
                                        st.info("Insufficient subject data points for trend analysis.")
                                else:
                                    st.info("No valid subject scores found for trend analysis.")
                        else:
                            st.info("Not enough valid data points to show progress trend.")
                    else:
                        st.info("Only one period of data available for this student.")
                else:
                    st.info("Period information not available in the data.")
                
                st.markdown("#### Detailed Records")
                detailed_df = student_data.loc[:, ~student_data.columns.str.contains('^Unnamed')]
                st.dataframe(detailed_df, use_container_width=True)

# --- Tab 3 ---
with tab3:
    st.markdown("### 📋 Detailed Student Data")
    
    columns_to_remove = [
        'Unnamed: 0_x', 'Unnamed: 18', 'Unnamed: 20', 'Woodwork', 'M %', 'MM/MP', 
        'Guardian', 'Contact', 'Unnamed: 6', 'Unnamed: 0_y', 'Unnamed: 9', 
        'Unnamed: 10', 'Unnamed: 11'
    ]
    
    display_df = df_main.copy()
    
    existing_unwanted_cols = [col for col in columns_to_remove if col in display_df.columns]
    if existing_unwanted_cols:
        display_df = display_df.drop(columns=existing_unwanted_cols)
    
    if len(display_df.columns) > 0:
        first_col = display_df.columns[0]
        if 'Unnamed' in str(first_col) or first_col == 0:
            display_df = display_df.drop(columns=[first_col])
    
    seen_columns = set()
    columns_to_keep = []
    columns_to_drop = []
    
    for col in display_df.columns:
        col_lower = str(col).lower().strip()
        if 'business' in col_lower and 'studies' in col_lower:
            if 'business_studies' not in seen_columns:
                seen_columns.add('business_studies')
                columns_to_keep.append(col)
            else:
                columns_to_drop.append(col)
        else:
            if col_lower not in seen_columns:
                seen_columns.add(col_lower)
                columns_to_keep.append(col)
            else:
                columns_to_drop.append(col)
    
    if columns_to_drop:
        display_df = display_df.drop(columns=columns_to_drop)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Records", len(display_df))
    with col2:
        if "M%" in display_df.columns:
            avg_performance = display_df["M%"].mean()
            st.metric("Average Performance", f"{avg_performance:.1f}%")
    with col3:
        if "School" in display_df.columns:
            unique_schools = display_df["School"].nunique()
            st.metric("Schools Represented", unique_schools)
    
    st.markdown("---")
    
    search_term = st.text_input("🔍 Search in data (student name, school, etc.)", "")
    if search_term:
        text_columns = display_df.select_dtypes(include=['object']).columns
        mask = False
        for col in text_columns:
            mask |= display_df[col].astype(str).str.contains(search_term, case=False, na=False)
        display_df = display_df[mask]
        st.info(f"Found {len(display_df)} records matching '{search_term}'")
    
    st.dataframe(display_df, use_container_width=True, height=600)
    
    if st.button("📥 Download Filtered Data as CSV"):
        csv = display_df.to_csv(index=False)
        st.download_button(
            label="Download CSV",
            data=csv,
            file_name=f"student_data_filtered_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )

# --- Tab 4 ---
with tab4:
    st.markdown("### 🚪 Dropouts Tracking")
    if dropout_df is not None and not dropout_df.empty:
        df = dropout_df.copy()
        header_row_idx = None
        for idx, row in df.iterrows():
            if 'Student Name' in row.values and 'Dropout Period' in row.values and 'Reason' in row.values:
                header_row_idx = idx
                break
        if header_row_idx is not None:
            df.columns = df.iloc[header_row_idx]
            df = df.iloc[header_row_idx+1:]
            df = df.reset_index(drop=True)
        needed_cols = ["Student Name", "Dropout Period", "Reason"]
        df = df[[col for col in needed_cols if col in df.columns]]
        df = df[df["Student Name"].astype(str).str.strip() != ""]
        df = df[df["Reason"].astype(str).str.strip() != ""]
        if "Dropout Period" in df.columns:
            df["Dropout Period"] = pd.to_datetime(df["Dropout Period"], errors='coerce').dt.strftime('%b-%y')
        st.dataframe(df, use_container_width=True)
        csv_data = df.to_csv(index=False)
        st.download_button(
            label="📥 Download Dropouts Data as CSV",
            data=csv_data,
            file_name=f"dropouts_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
    else:
        st.warning("No dropout data found or the file is empty.")

# --- Tab 5 ---
with tab5:
    if tertiary_df is not None and not tertiary_df.empty:
        
        # ==========================================
        # 1. TOP SECTION: Horizontal Filters
        # ==========================================
        st.markdown("""
            <div style='background-color: #87cefa; padding: 6px; border-radius: 4px; text-align: center; margin-bottom: 15px;'>
                <span style='color: black; font-weight: bold; font-size: 16px;'>🔍 Filter Tertiary Data</span>
            </div>
        """, unsafe_allow_html=True)

        def filter_df(exclude=None):
            temp = tertiary_df.copy()

            if exclude != "School" and st.session_state.get("t_school", "All") != "All":
                temp = temp[temp["School"] == st.session_state.t_school]

            if exclude != "Course" and st.session_state.get("t_course", "All") != "All":
                temp = temp[temp["Course"] == st.session_state.t_course]

            if exclude != "Student" and st.session_state.get("t_student", "All") != "All":
                temp = temp[temp["Student"] == st.session_state.t_student]

            if exclude != "Year" and st.session_state.get("t_year", "All") != "All":
                temp = temp[temp["Year"] == st.session_state.t_year]

            #if exclude != "Semester" and st.session_state.get("t_sem", "All") != "All":
            #    temp = temp[temp["Semester"] == st.session_state.t_sem]

            return temp


        f_col1, f_col2, f_col3, f_col4 = st.columns(4)

        with f_col1:
            school_options = ["All"] + sorted(
                filter_df(exclude="School")["School"].dropna().astype(str).unique()
            )
            selected_school = st.selectbox(
                "School",
                school_options,
                key="t_school"
            )

        with f_col2:
            course_options = ["All"] + sorted(
                filter_df(exclude="Course")["Course"].dropna().astype(str).unique()
            )
            selected_course = st.selectbox(
                "Course / Program",
                course_options,
                key="t_course"
            )

        with f_col3:
            student_options = ["All"] + sorted(
                filter_df(exclude="Student")["Student"].dropna().astype(str).unique()
            )
            selected_student = st.selectbox(
                "Student Name",
                student_options,
                key="t_student"
            )

        with f_col4:
            year_options = ["All"] + sorted(
                filter_df(exclude="Year")["Year"].dropna().astype(str).unique()
            )
            selected_year = st.selectbox(
                "Year",
                year_options,
                key="t_year"
            )

        # with f_col5:
            #semester_options = ["All"] + sorted(
            #    filter_df(exclude="Semester")["Semester"].dropna().astype(str).unique()
            #)
            #selected_semester = st.selectbox(
            #    "Semester",
            #    semester_options,
            #    key="t_sem"
            #)

        t_filtered = filter_df()

        st.markdown("---")

        # ==========================================
        # 2. METRICS SECTION (Full Width - 4 Metrics)
        # ==========================================
        # Smarter categorization logic using the 'Result' column generated during parsing
        def categorize_status(row):
            res = str(row.get('Result', '')).upper().strip()
            if res == 'PASS':
                return 'Passed'
            elif res == 'FAIL':
                return 'Failed'
            elif res in ['MISSING', 'SPECIAL']:
                return 'Pending'
            
        t_filtered['Status_Category'] = t_filtered.apply(categorize_status, axis=1)

        total_units = len(t_filtered)
        passed_count = len(t_filtered[t_filtered['Status_Category'] == 'Passed'])
        failed_count = len(t_filtered[t_filtered['Status_Category'] == 'Failed'])
        pending_count = len(t_filtered[t_filtered['Status_Category'] == 'Pending'])
        
        t_col1, t_col2, t_col3, t_col4 = st.columns(4)
        with t_col1:
            st.markdown('<div class="metric-header">Total Units Taken</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="metric-card"><div class="metric-value">{total_units}</div></div>', unsafe_allow_html=True)
        with t_col2:
            st.markdown('<div class="metric-header" style="background-color: #d4edda;">Units Passed</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#155724;">{passed_count}</div></div>', unsafe_allow_html=True)
        with t_col3:
            st.markdown('<div class="metric-header" style="background-color: #f8d7da;">Units Failed</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#721c24;">{failed_count}</div></div>', unsafe_allow_html=True)
        with t_col4:
            st.markdown('<div class="metric-header" style="background-color: #fff3cd;">Units Missing/Special</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="metric-card"><div class="metric-value" style="color:#856404;">{pending_count}</div></div>', unsafe_allow_html=True)

        st.markdown("---")
        
        # ==========================================
        # 3. CHARTS SECTION (Full Width)
        # ==========================================
        chart_col1, chart_col2 = st.columns(2)
        
        if "Grade" in t_filtered.columns:
            chart_data = t_filtered.dropna(subset=['Grade']).copy()
            if not chart_data.empty:
                grade_order = ["A", "B", "C", "D", "E", "F", "PASS", "FAIL", "NOT AVAILABLE", "NOT APPEARED"]
                fig_dist = px.histogram(
                    chart_data, x="Year", color="Grade", 
                    title="Grade Distribution by Year", barmode="group",
                    category_orders={"Grade": grade_order},
                    color_discrete_sequence=px.colors.qualitative.Pastel
                )
                fig_dist.update_layout(xaxis_title="Year", yaxis_title="Number of Units", height=450)
                chart_col1.plotly_chart(fig_dist, use_container_width=True)
            else:
                chart_col1.info("No graded units available for distribution chart based on current filters.")

            status_counts = t_filtered['Status_Category'].value_counts()
            
            fig_pie = px.pie(
                values=status_counts.values,
                names=status_counts.index,
                title="Overall Unit Status Breakdown",
                color=status_counts.index,
                color_discrete_map={"Passed": "#66b3ff", "Failed": "#ff9999", "Pending": "#ffcc99"}
            )
            fig_pie.update_layout(height=450)
            chart_col2.plotly_chart(fig_pie, use_container_width=True)
        
        st.markdown("---")
        
        # ==========================================
        # 4. DATA TABLES SECTION (Separated)
        # ==========================================
        full_display_cols = [col for col in ['Student', 'School', 'Course', 'Year', 'Semester', 'Code', 'Unit Name', 'Grade', 'Result'] if col in t_filtered.columns]
        table_config = {
            "School": st.column_config.TextColumn("School", width="medium"), 
            "Course": st.column_config.TextColumn("Course", width="medium"), 
            "Unit Name": st.column_config.TextColumn("Unit Name", width="medium")
        }

        # --- TABLE 1: MISSING / PENDING ---
        st.markdown("#### ⏳ Missing/Special Units")
        pending_df = t_filtered[t_filtered['Status_Category'] == 'Pending'].sort_values(by='Result', ascending=False)
        if not pending_df.empty:
            st.warning(f"Found {len(pending_df)} units where results are not available.These may require follow-up.")
            st.dataframe(pending_df[full_display_cols], use_container_width=True, hide_index=True, height=250, column_config=table_config)
        else:
            st.success("All grades are accounted for! 🎉")
            
        st.markdown("<br>", unsafe_allow_html=True)
        
        # --- TABLE 2: FAILED ---
        st.markdown("#### Failed Units ")
        failed_df = t_filtered[t_filtered['Status_Category'] == 'Failed']
        if not failed_df.empty:
            st.error(f"Found {len(failed_df)} failed units that may require retakes.")
            st.dataframe(failed_df[full_display_cols], use_container_width=True, hide_index=True, height=250, column_config=table_config)
        else:
            st.success("No failed units detected! 🎉")
            
        st.markdown("<br>", unsafe_allow_html=True)
        
        # --- TABLE 3: PASSED ---
        st.markdown("#### ✅ Passed Units")
        passed_df = t_filtered[t_filtered['Status_Category'] == 'Passed']
        if not passed_df.empty:
            st.dataframe(passed_df[full_display_cols], use_container_width=True, hide_index=True, height=400, column_config=table_config)
        else:
            st.info("No passed units to display based on current filters.")

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("📥 Download Filtered Tertiary Data as CSV"):
            csv_tertiary = t_filtered[full_display_cols].to_csv(index=False)
            st.download_button(label="Download CSV", data=csv_tertiary, file_name=f"tertiary_data_filtered_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv", mime="text/csv")
    else:
        st.warning("⚠️ No Tertiary Data found. Please check your data source connection.")
